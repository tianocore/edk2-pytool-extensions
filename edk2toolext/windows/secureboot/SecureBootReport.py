import argparse
import logging
import sys
import ctypes
import os
import json

# Some of these libraries are windows only, so we need to import them conditionally
if sys.platform == "win32":
    import pywintypes
    import win32api
    import win32process
    import win32security

# import this from edk2toollib so we can parse the dbx
from edk2toollib.uefi.authenticated_variables_structure_support import (
    EfiSignatureDatabase,
)

KERNEL32 = ctypes.windll.kernel32
EFI_VAR_MAX_BUFFER_SIZE = 1024 * 1024
DEFAULT_OUTPUT_FOLDER = os.path.join(".", "SecureBootFiles")

# The supported files to retrieve from UEFI and their guids, additional files can be added here
SECUREBOOT_FILES = {
    "dbx": "{d719b2cb-3d3a-4596-a3bc-dad00e67656f}"  # EFI_IMAGE_SECURITY_DATABASE_GUID
}

KNOWN_CERTIFICATES = [
    "11FB1A56B8CAE09CF7B06AD620A9F13C604D108BDAF443C115FE5848805241E5",  # "Canonical Certificate",
    "F22246C9E65BB4543CB892B6507C062196A788E5A190D2D5A877B94DFDFB2935",  # "Debian Certificate",
    #     "flathash_sha256": "20656b9c8cfd47aee8f71025dbaa73883d985f33871b3e5d2f65fdcbcb6a2b52",
    #     "component": "Cisco Sub Ca Certificate"
]

logging.basicConfig()
logger = logging.getLogger()
logger.setLevel(logging.INFO)

###################################################################################################
# Helper functions
###################################################################################################


def write_json_file(data, output_file):
    """Writes a JSON file.

    :param data: Data to write
    :param output_file: Output file to write to

    :return: None
    """
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    with open(output_file, "w") as f:
        f.write(json.dumps(data, indent=4))

    logger.info("Wrote report to %s", output_file)


def write_xlsx_file(report, output_file):
    """Writes an XLSX file.

    :param data: Data to write
    :param output_file: Output file to write to

    :return: None
    """
    try:
        import xlsxwriter
    except ImportError:
        raise ImportError("You must install xlsxwriter to use this function")

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    workbook = xlsxwriter.Workbook(output_file)

    for page in report.keys():
        worksheet = workbook.add_worksheet(page)

        if "dict" in report[page].keys():
            # Write the header
            col = 0
            row = 0
            headers = ["authenticode_hash"]

            # Build up the field headers in the dictionary
            for hash in report[page]["dict"]:
                for field in report[page]["dict"][hash]:
                    if field not in headers:
                        headers.append(field)
                # Only need to iterate through one hash
                break

            # write the headers
            worksheet.write_row(row, col, headers)
            row += 1

            for hash in report[page]["dict"]:
                # Write the hash
                worksheet.write(row, col, hash)
                col += 1

                # Write the fields
                for field in headers[1:]:  # Skip the hash
                    if field not in report[page]["dict"][hash]:
                        col += 1
                        continue

                    data = report[page]["dict"][hash][field]
                    if isinstance(data, list):
                        data = ", ".join(data)

                    worksheet.write(row, col, data)
                    col += 1

                col = 0
                row += 1

        elif "list" in report[page]:
            row = 0
            headers = ["authenticode_hash"]
            worksheet.write_row(row, col, headers)
            row += 1

            for hash in report[page]["list"]:
                worksheet.write(row, col, hash)
                row += 1

    workbook.close()

    logger.info("Wrote report to %s", output_file)


def convert_uefi_org_revocation_file_to_dict(file):
    """Converts the excel file to json."""

    try:
        import openpyxl
    except ImportError:
        logger.error("Please install openpyxl to use this feature")
        sys.exit(1)

    data = {}

    wb = openpyxl.load_workbook(file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Set up the variables in a way that is human readable
        # flat hash is useless
        convert_arch = {"64-bit": "x86_64", "32-bit": "x86", "64-bit ARM": "arm64"}

        authenticode_hash = row[1].upper()

        # if the hash isn't in the known certificates list than it's authenticode
        # the only two types are authenticode and certificate however in practice only authenticode is used
        meta_data = {
            "flat_hash_sha256": row[0],
            "component": row[2],
            "arch": convert_arch.get(row[3], row[3]),
            "partner": row[4],
            "type": "certificate"
            if authenticode_hash in KNOWN_CERTIFICATES
            else "authenticode",
            "cves": row[5],
            "date": row[6],
            "authority": None,
            "links": [],
        }

        # Add links to the cve's if they exist
        for cve in meta_data["cves"].split("; "):
            if cve == "" or "XXXXX" in cve or "N/A" in cve:
                break

            meta_data["links"].append(f"https://nvd.nist.gov/vuln/detail/{cve}")

        # Add the authority if it exists
        # See https://oofhours.com/2021/01/19/uefi-secure-boot-who-controls-what-can-run/
        if meta_data["partner"] != "Unknown":
            if meta_data["partner"] == "Microsoft":
                meta_data["authority"] = "Microsoft Windows Production PCA 2011"
            else:
                meta_data["authority"] = "Microsoft Corporation UEFI CA 2011"

        data[authenticode_hash] = meta_data

    return data


def convert_uefi_org_file(args):
    """
    Converts the excel file to json.

    """
    metadata = convert_uefi_org_revocation_file_to_dict(args.uefi_org_excel_file)

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    with open(args.output, "w") as f:
        f.write(json.dumps(metadata, indent=4))

    logger.info("Wrote report to %s", args.output)


###################################################################################################
# Classes
###################################################################################################


class FirmwareVariables(object):
    def __init__(self):
        # enable required SeSystemEnvironmentPrivilege privilege
        privilege = win32security.LookupPrivilegeValue(
            None, "SeSystemEnvironmentPrivilege"
        )

        token = win32security.OpenProcessToken(
            win32process.GetCurrentProcess(),
            win32security.TOKEN_READ | win32security.TOKEN_ADJUST_PRIVILEGES,
        )

        win32security.AdjustTokenPrivileges(
            token, False, [(privilege, win32security.SE_PRIVILEGE_ENABLED)]
        )

        win32api.CloseHandle(token)

        try:
            self._GetFirmwareEnvironmentVariable = (
                KERNEL32.GetFirmwareEnvironmentVariableW
            )
            self._GetFirmwareEnvironmentVariable.restype = ctypes.c_int
            self._GetFirmwareEnvironmentVariable.argtypes = [
                ctypes.c_wchar_p,
                ctypes.c_wchar_p,
                ctypes.c_void_p,
                ctypes.c_int,
            ]
        except AttributeError:
            self._GetFirmwareEnvironmentVariable = None
            logger.warning("Get function doesn't exist")

    def get_variable(self, name, guid):
        """Gets a firmware variable.

        Args:
            name (str): Name of the variable to get
            guid (str): GUID of the variable to get

        Returns:
          The value of the variable
        """
        if self._GetFirmwareEnvironmentVariable is None:
            raise NotImplementedError(
                "GetFirmwareEnvironmentVariable is not implemented"
            )

        buffer = ctypes.create_string_buffer(EFI_VAR_MAX_BUFFER_SIZE)
        buffer_size = ctypes.c_int(EFI_VAR_MAX_BUFFER_SIZE)
        result = self._GetFirmwareEnvironmentVariable(name, guid, buffer, buffer_size)
        if result == 0:
            last_error = win32api.GetLastError()

            raise pywintypes.error(
                last_error,
                "GetFirmwareEnvironmentVariable",
                win32api.FormatMessage(last_error),
            )

        return buffer.raw[:result]


###################################################################################################
# main Functions
###################################################################################################


def get_secureboot_files(args):
    """Gets the DBX from the system.

    Args:
        ArgumentParser (ArgumentParser.namespace): the namespace from the ArgumentParser

    Returns:
        0 on success
    """

    if sys.platform != "win32":
        raise NotImplementedError("This function is only implemented for Windows")

    os.makedirs(args.output, exist_ok=True)

    # To make this cross platform, we need to make FirmwareVariables support other platforms
    fw_vars = FirmwareVariables()

    var = fw_vars.get_variable(
        args.secureboot_file, SECUREBOOT_FILES[args.secureboot_file]
    )

    output_file = os.path.join(args.output, f"{args.secureboot_file}.bin")

    with open(output_file, "wb") as f:
        f.write(var)

    logger.info("Wrote %s", output_file)

    return 0


def parse_dbx(args):
    """Parses the DBX.

    Args:
        ArgumentParser (ArgumentParser.namespace): the namespace from the ArgumentParser

    Returns:
        0 on success
    """

    with open(args.revocations_file, "r") as rev_fs:
        revocations = json.loads(rev_fs.read())

        report = {
            "identified": {
                "dict": {},
                "total": 0,
                "note": "Represents all the hashes found in a systems dbx that match a provided revocation",
            },
            "missing_protections": {
                "dict": {},
                "total": 0,
                "note": "The remaining hashes in the provided revocation list that were not found in the system dbx",
            },
            "not_found": {
                "list": [],
                "total": 0,
                "note": "The hashes that were found in the dbx that are not in the provided revocation list",
            },
        }

        # Filter the revocations by arch if requested
        if args.filter_by_arch:
            filter_revocations = {}
            for revocation in revocations:
                if "arch" in revocations[revocation]:
                    if revocations[revocation]["arch"] == args.filter_by_arch:
                        filter_revocations[revocation] = revocations[revocation]
            revocations = filter_revocations

        with open(args.dbx_file, "rb") as dbx_fs:
            dbx = EfiSignatureDatabase(dbx_fs)
            for Esl in dbx.EslList:
                for signature in Esl.SignatureData_List:
                    formatted_signature = ""
                    for byte in signature.SignatureData:
                        formatted_signature += f"{byte:02X}"

                    # Is this in our revocation list?
                    if formatted_signature in revocations:
                        report["identified"]["dict"][formatted_signature] = revocations[
                            formatted_signature
                        ]
                        report["identified"]["dict"][formatted_signature]["count"] = 1

                        # if we have identified the revocation, remove it from the list
                        del revocations[formatted_signature]
                    # have we already seen this signature?
                    elif formatted_signature in report["identified"]["dict"]:
                        report["identified"]["dict"][formatted_signature]["count"] += 1
                    # otherwise, we have a signature that is not in our revocation list
                    else:
                        report["not_found"]["list"].append(formatted_signature)

        # Anything left in the revocation list is missing from the dbx
        for revocation in revocations:
            report["missing_protections"]["dict"][revocation] = revocations[revocation]

        # Add totals
        report["identified"]["total"] = len(report["identified"]["dict"])
        report["not_found"]["total"] = len(report["not_found"]["list"])
        report["missing_protections"]["total"] = len(
            report["missing_protections"]["dict"]
        )

        # Check for flat hashes
        for supposed_authenticode_hash in report["not_found"]["list"]:
            for supposed_flat_hash in revocations:
                if "flat_hash_sha256" not in revocations[supposed_flat_hash]:
                    break

                if supposed_authenticode_hash == revocations[supposed_flat_hash]["flat_hash_sha256"]:
                    logger.warning(
                        "This hash appears to be a flat sha256 hash: %s", supposed_authenticode_hash
                    )

    if args.format == "json":
        write_json_file(report, args.output + ".json")
    elif args.format == "xlsx":
        write_xlsx_file(report, args.output + ".xlsx")

    return 0


###################################################################################################
# Command Line Parsing Functions
###################################################################################################


def setup_parse_dbx(subparsers):
    """Setup the parse_dbx subparser.

    Args:
        subparsers (ArgumentParser): the subparsers object from the ArgumentParser

    Returns:
        subparsers

    """
    parser = subparsers.add_parser("parse_dbx")
    parser.set_defaults(function=parse_dbx)

    parser.add_argument("dbx_file", help="Input file to read the DBX from")

    parser.add_argument(
        "revocations_file",
        help="JSON file containing the revocations to check against the DBX",
    )

    parser.add_argument(
        "--filter-by-arch",
        help="Filter the output to only include the specified arch type",
        choices=["x86", "x86_64", "arm", "arm64"],
        default=None,
    )

    parser.add_argument(
        "--output",
        help="Output file to write the dbx contents to (note: extension will be based on the format)",
        default=os.path.join(DEFAULT_OUTPUT_FOLDER, "dbx_report"),
    )

    parser.add_argument(
        "--format",
        help="Format of the revocations file",
        choices=["json", "xlsx"],
        default="json",
    )

    return subparsers


def setup_parse_uefi_org_files(subparsers):
    """Setup the parse_uefi_org_files subparser.

    Args:
        subparsers (ArgumentParser): the subparsers object from the ArgumentParser

    Returns:
        subparsers
    """

    parser = subparsers.add_parser("convert_file")
    parser.set_defaults(function=convert_uefi_org_file)

    parser.add_argument(
        "uefi_org_excel_file",
        help="The excel file from uefi.org to parse, this will provide metadata about the signatures",
    )

    parser.add_argument(
        "--output",
        help="Output file to write the uefi org revocations to",
        default=os.path.join(DEFAULT_OUTPUT_FOLDER, "uefi_org_revocations.json"),
    )

    return subparsers


def setup_get_secureboot_files(subparsers):
    """Setup the get_dbx subparser.

    Args:
        subparsers (ArgumentParser): the subparsers object from the ArgumentParser

    Returns:
        subparsers

    """
    parser = subparsers.add_parser("get")
    parser.set_defaults(function=get_secureboot_files)

    parser.add_argument(
        "secureboot_file",
        help="The secureboot file to get from UEFI",
        choices=SECUREBOOT_FILES.keys(),
    )

    parser.add_argument(
        "--output",
        help="Output folder to write the secureboot variables to",
        default=DEFAULT_OUTPUT_FOLDER,
    )

    return subparsers


def parse_args():
    """Parses arguments from the command line."""
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers()
    subparsers = setup_get_secureboot_files(subparsers)
    subparsers = setup_parse_uefi_org_files(subparsers)
    subparsers = setup_parse_dbx(subparsers)

    args = parser.parse_args()

    if not hasattr(args, "function"):
        parser.print_help(sys.stderr)
        sys.exit(1)

    return args


def main():
    """Main function."""
    args = parse_args()

    # Return the exit code from the function
    return sys.exit(args.function(args))


if __name__ == "__main__":
    main()
