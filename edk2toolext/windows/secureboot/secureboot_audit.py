# @file This file tests secureboot_audit.py script
# Command-line tool for inspecting UEFI Secure Boot databases
# Requires "pip install edk2-pytool-extensions"
##
# Copyright (c) Microsoft Corporation
#
# SPDX-License-Identifier: BSD-2-Clause-Patent
##
"""Command-line tool for inspecting UEFI Secure Boot databases."""

import argparse
import csv
import ctypes
import json
import logging
import os
import sys
from typing import BinaryIO

import openpyxl
import xlsxwriter

# Some of these libraries are windows only, so we need to import them conditionally
if sys.platform == "win32":
    import pywintypes
    import win32api
    import win32process
    import win32security

# import this from edk2toollib so we can parse the dbx
from edk2toollib.uefi.authenticated_variables_structure_support import (
    EfiSignatureDatabase,
)

logging.basicConfig()
logger = logging.getLogger()
logger.setLevel(logging.INFO)

if sys.platform == "win32":
    KERNEL32 = ctypes.windll.kernel32
    EFI_VAR_MAX_BUFFER_SIZE = 1024 * 1024

DEFAULT_OUTPUT_FOLDER = os.path.join(".", "SecureBootFiles")

# The supported files to retrieve from UEFI and their guids, additional files can be added here
SECUREBOOT_FILES = {
    "dbx": "{d719b2cb-3d3a-4596-a3bc-dad00e67656f}",  # EFI_IMAGE_SECURITY_DATABASE_GUID
    "db": "{d719b2cb-3d3a-4596-a3bc-dad00e67656f}",  # EFI_IMAGE_SECURITY_DATABASE_GUID
    "dbt": "{d719b2cb-3d3a-4596-a3bc-dad00e67656f}",  # EFI_IMAGE_SECURITY_DATABASE_GUID
    "KEK": "{8BE4DF61-93CA-11d2-AA0D-00E098032B8C}",  # EFI_GLOBAL_VARIABLE
    "PK": "{8BE4DF61-93CA-11d2-AA0D-00E098032B8C}",  # EFI_GLOBAL_VARIABLE
}

KNOWN_CERTIFICATES = [
    "11FB1A56B8CAE09CF7B06AD620A9F13C604D108BDAF443C115FE5848805241E5",  # "Canonical Certificate",
    "F22246C9E65BB4543CB892B6507C062196A788E5A190D2D5A877B94DFDFB2935",  # "Debian Certificate",
]


###################################################################################################
# Helper functions
###################################################################################################


def write_json_file(report: dict, output_file: str) -> None:
    """Writes a JSON file.

    Args:
        report (dict): report to write
        output_file (str): Output file to write to

    Returns:
        None
    """
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(json.dumps(report, indent=4))

    logger.info("Wrote report to %s", output_file)


def write_xlsx_file(report: dict, output_file: str) -> None:
    """Writes an XLSX file.

    Args:
        report (dict): Data to write
        output_file (str): Output file to write to

    Returns:
        None
    """
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    workbook = xlsxwriter.Workbook(output_file)

    for page in report.keys():
        worksheet = workbook.add_worksheet(page)

        if "dict" in report[page].keys():
            # Write the header
            col = 0
            row = 0
            headers = ["authenticode_hash"]

            # Build up the field headers in the dictionary
            for auth_hash in report[page]["dict"]:
                for field in report[page]["dict"][auth_hash]:
                    if field not in headers:
                        headers.append(field)
                # Only need to iterate through one hash
                break

            # write the headers
            worksheet.write_row(row, col, headers)
            row += 1

            for auth_hash in report[page]["dict"]:
                # Write the hash
                worksheet.write(row, col, auth_hash)
                col += 1

                # Write the fields
                for field in headers[1:]:  # Skip the hash
                    if field not in report[page]["dict"][auth_hash]:
                        col += 1
                        continue

                    data = report[page]["dict"][auth_hash][field]
                    if isinstance(data, list):
                        data = ", ".join(data)

                    worksheet.write(row, col, data)
                    col += 1

                col = 0
                row += 1

        elif "list" in report[page]:
            row = 0
            headers = ["authenticode_hash"]
            worksheet.write_row(row, col, headers)
            row += 1

            for auth_hash in report[page]["list"]:
                worksheet.write(row, col, auth_hash)
                row += 1

    workbook.close()

    logger.info("Wrote report to %s", output_file)


def convert_row_to_metadata(row: list) -> dict:
    """Converts a row from the csv to a metadata dictionary.

    Args:
        row (list): A row from the csv file

    Returns:
        dict: A dictionary containing the metadata
    """
    convert_arch = {"64-bit": "x86_64", "32-bit": "x86", "64-bit ARM": "arm64"}
    authenticode_hash = row[1].upper()

    def map_cve(cve: str) -> str:
        """Maps a CVE to the correct format.

        Args:
            cve (str): CVE to map

        Returns:
            str: Mapped CVE
        """
        # Intentionally naive mapping of CVE's to the correct format
        if "CVE" in cve.upper():
            # this condition could have multiple CVE's in it
            return cve
        elif "black lotus" in cve.lower():
            return "CVE-2022-21894"
        else:
            return "N/A"

    # if the hash isn't in the known certificates list than it's authenticode
    # the only two types are authenticode and certificate however in practice only authenticode is used
    meta_data = {
        "flat_hash_sha256": row[0],
        "component": row[2] if row[2] != "" else None,
        "arch": convert_arch.get(row[3], row[3]),
        "partner": row[4],
        "type": "certificate" if authenticode_hash in KNOWN_CERTIFICATES else "authenticode",
        "cves": map_cve(row[5]),
        "date": row[6].replace("\n", ""),
        "authority": None,
        "links": [],
    }

    # Add links to the cve's if they exist
    for cve in meta_data["cves"].split("; "):
        if cve == "" or "XXXXX" in cve or "N/A" in cve:
            break

        meta_data["links"].append(f"https://nvd.nist.gov/vuln/detail/{cve}")

    # Add the authority if it exists
    # See https://oofhours.com/2021/01/19/uefi-secure-boot-who-controls-what-can-run/
    if meta_data["partner"] != "Unknown":
        if meta_data["partner"] == "Microsoft":
            meta_data["authority"] = "Microsoft Windows Production PCA 2011"
        else:
            meta_data["authority"] = "Microsoft Corporation UEFI CA 2011"

    return authenticode_hash, meta_data


def convert_uefi_org_revocation_file_to_dict(file: str) -> dict:
    """Converts the excel file to json.

    Args:
        file (str): Path to the excel or csv file

    Returns:
        dict: The data from the excel or csv file
    """
    data = {}

    if file.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            authenticode_hash, meta_data = convert_row_to_metadata(row)
            data[authenticode_hash] = meta_data
    elif file.endswith(".csv"):
        with open(file, "r") as f:
            for i, row in enumerate(csv.reader(f)):
                if i == 0:
                    # Skip the header
                    continue
                authenticode_hash, meta_data = convert_row_to_metadata(row)
                data[authenticode_hash] = meta_data
    else:
        # This should never happen
        raise Exception("Unknown file type")

    return data


def convert_uefi_org_file(args: argparse.Namespace) -> None:
    """Converts the excel file to json.

    Args:
        args (argparse.Namespace): The arguments

    Returns:
        None
    """
    metadata = convert_uefi_org_revocation_file_to_dict(args.uefi_org_file)

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    with open(args.output, "w") as f:
        f.write(json.dumps(metadata, indent=4))

    logger.info("Wrote report to %s", args.output)


def generate_dbx_report(dbx_fs: BinaryIO, revocations: dict) -> dict:
    """Generates a report of the dbx.

    Args:
        dbx_fs (filestream): The dbx file system
        revocations (dict): The revocation list

    Returns:
        dict: The report
    """
    report = {
        "identified": {
            "dict": {},
            "total": 0,
            "note": "Represents all the hashes found in a systems dbx that match a provided revocation",
        },
        "missing_protections": {
            "dict": {},
            "total": 0,
            "note": "The remaining hashes in the provided revocation list that were not found in the system dbx",
        },
        "not_found": {
            "list": [],
            "total": 0,
            "note": "The hashes that were found in the dbx that are not in the provided revocation list",
        },
    }

    dbx = EfiSignatureDatabase(dbx_fs)
    for signature_list in dbx.EslList:
        for signature in signature_list.SignatureData_List:
            formatted_signature = ""
            for byte in signature.SignatureData:
                formatted_signature += f"{byte:02X}"

            # Is this in our revocation list?
            if formatted_signature in revocations:
                report["identified"]["dict"][formatted_signature] = revocations[formatted_signature]
                report["identified"]["dict"][formatted_signature]["count"] = 1

                # if we have identified the revocation, remove it from the list
                del revocations[formatted_signature]
            # have we already seen this signature?
            elif formatted_signature in report["identified"]["dict"]:
                report["identified"]["dict"][formatted_signature]["count"] += 1
            # otherwise, we have a signature that is not in our revocation list
            else:
                report["not_found"]["list"].append(formatted_signature)

    # Anything left in the revocation list is missing from the dbx
    for revocation in revocations:
        report["missing_protections"]["dict"][revocation] = revocations[revocation]

    # Add totals
    report["identified"]["total"] = len(report["identified"]["dict"])
    report["not_found"]["total"] = len(report["not_found"]["list"])
    report["missing_protections"]["total"] = len(report["missing_protections"]["dict"])

    # Check for flat hashes
    for supposed_authenticode_hash in report["not_found"]["list"]:
        for supposed_flat_hash in revocations:
            if "flat_hash_sha256" not in revocations[supposed_flat_hash]:
                break

            if supposed_authenticode_hash == revocations[supposed_flat_hash]["flat_hash_sha256"]:
                logger.warning("This hash appears to be a flat sha256 hash: %s", supposed_authenticode_hash)

    return report


def filter_revocation_list_by_arch(revocations: dict, filter_by_arch: str = None) -> dict:
    """Filters the revocation list by architecture.

    Args:
        revocations (dict): The revocation list
        filter_by_arch (str): The architecture to filter by

    Returns:
        dict: The filtered revocation list
    """
    # Filter the revocations by arch if requested
    if filter_by_arch:
        filter_revocations = {}
        for revocation in revocations:
            if "arch" in revocations[revocation]:
                if revocations[revocation]["arch"] == filter_by_arch:
                    filter_revocations[revocation] = revocations[revocation]
        revocations = filter_revocations

    return revocations


###################################################################################################
# Classes
###################################################################################################


class FirmwareVariables(object):
    """Class to interact with firmware variables."""

    def __init__(self) -> None:
        """Constructor."""
        # enable required SeSystemEnvironmentPrivilege privilege
        privilege = win32security.LookupPrivilegeValue(None, "SeSystemEnvironmentPrivilege")

        token = win32security.OpenProcessToken(
            win32process.GetCurrentProcess(),
            win32security.TOKEN_READ | win32security.TOKEN_ADJUST_PRIVILEGES,
        )

        win32security.AdjustTokenPrivileges(token, False, [(privilege, win32security.SE_PRIVILEGE_ENABLED)])

        win32api.CloseHandle(token)

        try:
            self._GetFirmwareEnvironmentVariable = KERNEL32.GetFirmwareEnvironmentVariableW
            self._GetFirmwareEnvironmentVariable.restype = ctypes.c_int
            self._GetFirmwareEnvironmentVariable.argtypes = [
                ctypes.c_wchar_p,
                ctypes.c_wchar_p,
                ctypes.c_void_p,
                ctypes.c_int,
            ]
        except AttributeError:
            self._GetFirmwareEnvironmentVariable = None
            logger.warning("Get function doesn't exist")

    def get_variable(self, name: str, guid: str) -> bytes:
        """Gets a firmware variable.

        Args:
            name (str): Name of the variable to get
            guid (str): GUID of the variable to get

        Returns:
          The value of the variable
        """
        if self._GetFirmwareEnvironmentVariable is None:
            raise NotImplementedError("GetFirmwareEnvironmentVariable is not implemented")

        buffer = ctypes.create_string_buffer(EFI_VAR_MAX_BUFFER_SIZE)
        buffer_size = ctypes.c_int(EFI_VAR_MAX_BUFFER_SIZE)
        result = self._GetFirmwareEnvironmentVariable(name, guid, buffer, buffer_size)
        if result == 0:
            last_error = win32api.GetLastError()

            raise pywintypes.error(
                last_error,
                "GetFirmwareEnvironmentVariable",
                win32api.FormatMessage(last_error),
            )

        return buffer.raw[:result]


###################################################################################################
# main Functions
###################################################################################################


def get_secureboot_files(args: argparse.Namespace) -> int:
    """Gets the DBX from the system.

    Args:
        args (argparse.Namespace): the namespace from the ArgumentParser

    Returns:
        0 on success
    """
    if sys.platform != "win32":
        raise NotImplementedError("This function is only implemented for Windows")

    os.makedirs(args.output, exist_ok=True)

    # To make this cross platform, we need to make FirmwareVariables support other platforms
    fw_vars = FirmwareVariables()

    var = fw_vars.get_variable(args.secureboot_file, SECUREBOOT_FILES[args.secureboot_file])

    output_file = os.path.join(args.output, f"{args.secureboot_file}.bin")

    with open(output_file, "wb") as f:
        f.write(var)

    with open(output_file, "rb") as f:
        EfiSignatureDatabase(f).Print()

    logger.info("Wrote %s", output_file)

    return 0


def parse_dbx(args: argparse.Namespace) -> int:
    """Parses the DBX.

    Args:
        args (argparse.Namespace): the namespace from the ArgumentParser

    Returns:
        0 on success
    """
    report = {}
    with open(args.revocations_file, "r", encoding="utf-8") as rev_fs:
        revocations = json.loads(rev_fs.read())

        with open(args.dbx_file, "rb") as dbx_fs:
            revocations = filter_revocation_list_by_arch(revocations, args.filter_by_arch)
            report = generate_dbx_report(dbx_fs, revocations)

    if args.format == "json":
        write_json_file(report, args.output + ".json")
    elif args.format == "xlsx":
        write_xlsx_file(report, args.output + ".xlsx")

    return 0


###################################################################################################
# Command Line Parsing Functions
###################################################################################################


def valid_file(param: str, valid_extensions: tuple = (".csv", ".xlsx")) -> str:
    """Checks if a file is valid.

    Args:
        param (str): the file to check
        valid_extensions (tuple): the valid extensions

    Returns:
        str: the file if it is valid
    """
    base, ext = os.path.splitext(param)
    if ext.lower() not in valid_extensions:
        raise argparse.ArgumentTypeError("File must be one of the following types: {}".format(valid_extensions))
    return param


def setup_parse_dbx(subparsers: argparse._SubParsersAction) -> argparse._SubParsersAction:
    """Setup the parse_dbx subparser.

    Args:
        subparsers (argparse._SubParsersAction): the subparsers object from the ArgumentParser

    Returns:
        subparsers (argparse._SubParsersAction): the subparsers object from the ArgumentParser
    """
    parser = subparsers.add_parser("parse_dbx")
    parser.set_defaults(function=parse_dbx)

    parser.add_argument("dbx_file", help="Input file to read the DBX from")

    parser.add_argument(
        "revocations_file",
        help="JSON file containing the revocations to check against the DBX",
    )

    parser.add_argument(
        "--filter-by-arch",
        help="Filter the output to only include the specified arch type",
        choices=["x86", "x86_64", "arm", "arm64"],
        default=None,
    )

    parser.add_argument(
        "--output",
        help="Output file to write the dbx contents to" + " (note: extension will be based on the format)",
        default=os.path.join(DEFAULT_OUTPUT_FOLDER, "dbx_report"),
    )

    parser.add_argument(
        "--format",
        help="Format of the revocations file",
        choices=["json", "xlsx"],
        default="json",
    )

    return subparsers


def setup_parse_uefi_org_files(subparsers: argparse._SubParsersAction) -> argparse._SubParsersAction:
    """Setup the parse_uefi_org_files subparser.

    Args:
        subparsers (argparse._SubParsersAction): the subparsers object from the ArgumentParser

    Returns:
        subparsers (argparse._SubParsersAction): the subparsers object from the ArgumentParser
    """
    parser = subparsers.add_parser("convert_file")
    parser.set_defaults(function=convert_uefi_org_file)

    parser.add_argument(
        "uefi_org_file",
        type=valid_file,
        help="The csv or excel file from uefi.org to parse, this will provide metadata about the signatures",
    )

    parser.add_argument(
        "--output",
        help="Output file to write the uefi org revocations to",
        default=os.path.join(DEFAULT_OUTPUT_FOLDER, "uefi_org_revocations.json"),
    )

    return subparsers


def setup_get_secureboot_files(subparsers: argparse._SubParsersAction) -> argparse._SubParsersAction:
    """Setup the get_dbx subparser.

    Args:
        subparsers (argparse._SubParsersAction): the subparsers object from the ArgumentParser

    Returns:
        subparsers (argparse._SubParsersAction): the subparsers object from the ArgumentParser
    """
    parser = subparsers.add_parser("get")
    parser.set_defaults(function=get_secureboot_files)

    parser.add_argument(
        "secureboot_file",
        help="The secureboot file to get from UEFI",
        choices=SECUREBOOT_FILES.keys(),
    )

    parser.add_argument(
        "--output",
        help="Output folder to write the secureboot variables to",
        default=DEFAULT_OUTPUT_FOLDER,
    )

    return subparsers


def parse_args() -> argparse.Namespace:
    """Parses arguments from the command line."""
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers()
    subparsers = setup_get_secureboot_files(subparsers)
    subparsers = setup_parse_uefi_org_files(subparsers)
    subparsers = setup_parse_dbx(subparsers)

    args = parser.parse_args()

    if not hasattr(args, "function"):
        parser.print_help(sys.stderr)
        sys.exit(1)

    return args


def main() -> None:
    """Main function."""
    args = parse_args()

    # Raise a SystemExit exception with the exit status
    sys.exit(args.function(args))


if __name__ == "__main__":
    main()
